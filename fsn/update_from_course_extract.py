from threading import Thread
import unicodedata
import openpyxl
import os
import re

from eliasmamo_import import *
from secret import TOKEN
from configuration import TEAM_ID

def main():
    driver = Driver(
            {
                'url': 'mattermost.fysiksektionen.se',
                'basepath': '/api/v4',
                'verify': True,
                'scheme': 'https',
                'port': 443,
                'auth': None,
                'token': TOKEN,
                'keepalive': True,
                'keepalive_delay': 5,
                }
            )

    driver.login()

    #### --- Config ----
    # filename = "VT24_p3.xlsx"
    filename = "Testing.xlsx"
    sheet_name = "Sheet1"
    program_column = 1
    kth_id_column = 2
    course_code_column = 6
    course_name_column = 7
    course_version_column = 8
    #### ---- Mattermost ----
    MAX_CHANNEL_URL_LENGTH = 64
    MAX_DISPLAY_NAME_LENGTH = 64

    dataframe = openpyxl.load_workbook(os.path.join(os.path.dirname(__file__), "kursutdrag", filename), read_only = True)

    if sheet_name not in dataframe.sheetnames:
        print(f"Could not find '{sheet_name}' among the sheets, available sheets are: {dataframe.sheetnames}.")
        return

    sheet = dataframe[sheet_name]

    print(f"Using columns with headers: ")
    print(f"    Program: {sheet.cell(row = 1, column = program_column).value}")
    print(f"    KTH id: {sheet.cell(row = 1, column = kth_id_column).value}")
    print(f"    Course code: {sheet.cell(row = 1, column = course_code_column).value}")
    print(f"    Course name: {sheet.cell(row = 1, column = course_name_column).value}")
    print(f"    Course version: {sheet.cell(row = 1, column = course_version_column).value}")
    confirmation = input("Please confirm this looks right? (y/N): ")
    if confirmation not in ["y", "Y", "yes", "Yes"]:
        return

    users = {}
    for user in get_all_users(driver):
        users[user["username"]] = user["id"]

    channels = {}
    for channel in get_all_private_channels(driver, TEAM_ID):
        if channel["creator_id"] == driver.client.userid:
            channels[channel["name"]] = channel["id"]

    # Validation
    row_index = 2
    validation_error = False
    for row in sheet.iter_rows(2):
        if re.match(r"^[A-Z]{2}[A-Z0-9]{4}$", str(row[course_code_column - 1].value)) is None:
            print(f"row {row_index}: Invalid format of course code. {str(row[course_code_column - 1].value)}")
            validation_error = True

        if re.match(r"^[0-9]{5}$", str(row[course_version_column - 1].value)) is None:
            print(f"row {row_index}: Invalid format of course version. {str(row[course_version_column - 1].value)}")
            validation_error = True

        row_index += 1

    if validation_error:
        ans = input(f"Encountered validation error, do you wish to continue (skipping lines with validation error)? (y/N): ")
        if ans not in ["y", "Y", "yes", "Yes"]:
            return

    skipped_kth_ids = set()

    row_index = 2
    for row in sheet.iter_rows(2):
        kth_mail = str(row[kth_id_column - 1].value)
        kth_id = kth_mail.replace("@kth.se", "")

        if kth_id in skipped_kth_ids:
            continue

        if kth_id not in users:
            print(f"{kth_id} does not have a mattermost account. Skipping...")
            skipped_kth_ids.add(kth_id)
            continue

        if kth_id not in ["ellundel", "eskilny", "wkraft"]: ## TODO: Debug to only affect ellundel@kth.se 
            continue

        if re.match(r"^[0-9]{5}$", str(row[course_version_column - 1].value)) is None or \
           re.match(r"^[A-Z]{2}[A-Z0-9]{4}$", str(row[course_code_column - 1].value)) is None:
               print(f"row {row_index} (kthid {kth_id}) contains validation error, skipping row...")
               continue

        course_code = str(row[course_code_column - 1].value)
        course_name = str(row[course_name_column - 1].value)
        course_version = str(row[course_version_column - 1].value)

        channel_name = f"{course_code.lower()}-{course_name.lower().replace(' ', '-')}"
        channel_name = unicodedata.normalize("NFKD", channel_name).encode("ascii", errors = "ignore").decode("ascii")
        channel_course_version = f"-{course_version}"
        channel_name = channel_name[:MAX_CHANNEL_URL_LENGTH - len(channel_course_version)] + channel_course_version

        # Clean unwanted characters and reduce length to maximum allowed
        channel_name = re.sub(r"[^a-z0-9-]", "", channel_name)[:MAX_CHANNEL_URL_LENGTH]

        if channel_name not in channels:
            ans = input(f"Missing channel {channel_name}, create a new channel? (y/N): ")
            if ans not in ["y", "Y", "yes", "Yes"]:
                print(f"Not creating a new channel. Skipping course for user {kth_id}...")
                continue
            print(f"Creating new channel {channel_name}")
            channel_course_version_display = f"({course_version})"
            channel_display_name = f"{course_code} {course_name} "[:MAX_DISPLAY_NAME_LENGTH - len(channel_course_version_display)] + channel_course_version_display
            new_channel = driver.channels.create_channel({"team_id": TEAM_ID, "name": channel_name, "display_name": channel_display_name, "type": "P"})
            channels[channel_name] = new_channel["id"]

            print(f"Sending information message in {channel_name}")
            driver.posts.create_post({
                "channel_id": new_channel["id"], 
                "message": f"""### Welcome to {course_code} {course_name} ({course_version})

##### This is a private channel where you can discuss this course. Only the current members of the course are part of this channel. If you leave this channel you can join it again by writing `join {course_code.lower()} {course_version.lower()}` in a DM to @fsn-bot.

In here you can use $\\LaTeX$ to write math formulas. For example:
```
$\\int_{{0}}^{{\\infty}}{{\\frac{{f(x)}}{{g(x)}}dx}}$
```
$\\int_{{0}}^{{\\infty}}{{\\frac{{f(x)}}{{g(x)}}dx}}$

*If you encounter any problems, require moderation, wish to not be added to new course channels in the future, or have any other type of question, please send an email to [mattermost@f.kth.se](mailto:mattermost@f.kth.se) or contact @ellundel or @eskilny by DM.*"""})

        if channel_name not in channels:
            continue


        with open(os.path.join(os.path.dirname(__file__), "added_to_channel", f"{course_code.lower()}-{course_version.lower()}.txt"), "r+") as f:
            for line in f:
                if users[kth_id] in line:
                    print(f"User {kth_id} has already been added to {channel_name}. Skipping...")
                    continue

        
        print(f"Adding user {kth_id} to channel {channel_name}...")
        driver.channels.add_user(channels[channel_name], {"user_id": users[kth_id]})
        with open(os.path.join(os.path.dirname(__file__), "added_to_channel", f"{course_code.lower()}-{course_version.lower()}.txt"), "a") as f:
            f.write(f"{users[kth_id]}\n")

        row_index += 1

    delete_new_posts_in_clean_channels(driver, channels)

if __name__ == "__main__":
    main()
